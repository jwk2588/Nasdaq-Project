import xmlschema
import openpyxl

# Load the US GAAP 2024 schema
schema = xmlschema.XMLSchema('schemas/us-gaap-2024.xsd')

# Check that the schema loaded correctly
print(f'Schema loaded: {schema.is_valid}')

# Load your Excel file from the network path
workbook = openpyxl.load_workbook(r'\\JWK3\Users\jklei\Documents\GM Financial Statments Transposed.xlsx')
sheet = workbook.active  # Access the active sheet

# Example: Reading a value from cell A1
cell_value = sheet['A1'].value
print(f'Value in A1: {cell_value}')

# Iterate over the Excel rows to map XBRL tags to financial data
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=8):
    broad_tag = row[0].value  # Column A (Broad Tags)
    detailed_tag = row[1].value  # Column B (Detailed Tags)
    financial_value = row[3].value  # Example: Column D (2020-12-31 data)
    
    # Validate the XBRL tag against the schema
    if detailed_tag in schema.elements:
        print(f"Tag '{detailed_tag}' is valid for value: {financial_value}")
        # Store this information in a validation result column (e.g., Column C)
        row[2].value = "Valid"
    else:
        print(f"Tag '{detailed_tag}' is not valid!")
        row[2].value = "Invalid"

# Save the updated Excel file with validation results
workbook.save(r'\\JWK3\Users\jklei\Documents\GM_Financial_Statements_Tagged.xlsx')
